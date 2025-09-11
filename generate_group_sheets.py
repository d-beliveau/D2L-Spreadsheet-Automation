import os
import re
from typing import Dict, List, Optional

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.utils.cell import coordinate_to_tuple


def _read_master(master_path: str) -> pd.DataFrame:
    """
    Read D2L master export. Supports CSV and XLSX.
    """
    _, ext = os.path.splitext(master_path.lower())
    if ext == ".csv":
        # D2L CSVs are usually UTF-8 or UTF-8 with BOM; try utf-8 first
        try:
            return pd.read_csv(master_path)
        except UnicodeDecodeError:
            return pd.read_csv(master_path, encoding="utf-8-sig")
    elif ext in (".xlsx", ".xls"):
        return pd.read_excel(master_path)
    else:
        raise ValueError(f"Unsupported file type: {ext}. Use .csv or .xlsx")


def _sanitize_filename(name: str) -> str:
    """
    Make a safe filename from a group or email string.
    """
    name = name.strip().replace("@", "_at_")
    # Remove characters invalid on Windows/*nix filesystems
    name = re.sub(r'[\\/*?:"<>|]+', "_", name)
    # Collapse whitespace
    name = re.sub(r"\s+", "_", name)
    return name[:200]  # keep it sane


def build_groups(
    df: pd.DataFrame,
    email_col: str = "Email",
    group_col: Optional[str] = None,
    name_cols: Optional[List[str]] = None,
) -> Dict[str, List[str]]:
    """
    Build mapping of {group_name -> list of member emails}.

    - If group_col is provided and present, groups by that column.
    - Otherwise returns individuals as one-person "groups" keyed by email.

    Args:
        df: DataFrame of D2L master.
        email_col: column containing student emails.
        group_col: column containing group names/IDs (optional).
        name_cols: optional list of columns to assist with filenames if no group_col.

    Returns:
        dict mapping group name -> list of emails
    """
    if email_col not in df.columns:
        raise ValueError(f"Email column '{email_col}' not found in master.")

    if group_col and group_col in df.columns:
        groups = {}
        for gval, sub in df.groupby(group_col):
            emails = [str(e).strip() for e in sub[email_col].dropna().astype(str).tolist()]
            if emails:
                groups[str(gval)] = emails
        return groups

    # Per-person fallback (no group column)
    groups = {}
    for _, row in df.iterrows():
        email = str(row[email_col]).strip()
        if not email:
            continue
        if name_cols and all(col in df.columns for col in name_cols):
            name_bits = [str(row[c]).strip() for c in name_cols]
            key = " ".join([b for b in name_bits if b]) or email
        else:
            key = email
        groups[key] = [email]
    return groups


def create_group_workbook(
    members: List[str],
    output_path: str,
    student_cell: str = "B4",
    grade_cell: str = "B11",
    template_path: Optional[str] = None,
    sheet_name: Optional[str] = None,
    overwrite: bool = False,
) -> str:
    """
    Create a single group workbook with members listed in student_cell.

    If template_path is provided, the template is loaded first; otherwise a new workbook is created.
    Leaves grade_cell blank for graders.

    Args:
        members: list of member emails.
        output_path: full path to write .xlsx
        student_cell: Excel cell (e.g., "B4") to write comma-separated member emails.
        grade_cell: Excel cell (e.g., "B11") reserved for total grade (left blank).
        template_path: optional .xlsx template to copy layout from.
        sheet_name: optional name of active sheet (if template has multiple sheets, pass the one to target).
        overwrite: overwrite existing file if True.

    Returns:
        The output_path written.
    """
    if not overwrite and os.path.exists(output_path):
        raise FileExistsError(f"File exists: {output_path}. Use overwrite=True to replace.")

    if template_path:
        wb = load_workbook(template_path)
        ws = wb[sheet_name] if (sheet_name and sheet_name in wb.sheetnames) else wb.active
    else:
        wb = Workbook()
        ws = wb.active
        if sheet_name:
            ws.title = sheet_name
        # Minimal helpful labels if no template supplied
        ws["A3"] = "Group Members (emails):"
        ws["A10"] = "Total Grade:"

    s_row, s_col = coordinate_to_tuple(student_cell)
    g_row, g_col = coordinate_to_tuple(grade_cell)

    # Write members as comma-separated list
    ws.cell(row=s_row, column=s_col, value=", ".join(members))

    # Ensure grade cell exists (leave blank)
    _ = ws.cell(row=g_row, column=g_col)

    # Ensure directory and save
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
    return output_path


def generate_group_sheets(
    master_path: str,
    output_folder: str,
    email_col: str = "Email",
    group_col: Optional[str] = None,
    name_cols: Optional[List[str]] = None,
    student_cell: str = "B4",
    grade_cell: str = "B11",
    template_path: Optional[str] = None,
    sheet_name: Optional[str] = None,
    overwrite: bool = False,
) -> List[str]:
    """
    High-level API: read master, build groups, and write one .xlsx per group/person.

    Returns:
        List of generated file paths.
    """
    df = _read_master(master_path)
    groups = build_groups(df, email_col=email_col, group_col=group_col, name_cols=name_cols)

    written = []
    for group_name, members in groups.items():
        fname = _sanitize_filename(group_name) + ".xlsx"
        out_path = os.path.join(output_folder, fname)
        path = create_group_workbook(
            members=members,
            output_path=out_path,
            student_cell=student_cell,
            grade_cell=grade_cell,
            template_path=template_path,
            sheet_name=sheet_name,
            overwrite=overwrite,
        )
        written.append(path)
    return written


# CLI
if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(description="Generate individual/group spreadsheets from a D2L master export.")
    parser.add_argument("master", help="Path to D2L master (.csv or .xlsx)")
    parser.add_argument("outdir", help="Output folder for generated .xlsx files")
    parser.add_argument("--email_col", default="Email", help="Column name for student email in master (default: Email)")
    parser.add_argument("--group_col", default=None, help="Column name for grouping (if omitted, generates per person)")
    parser.add_argument("--name_cols", nargs="*", default=None, help="Optional name columns to use for filenames if no group_col (e.g. --name_cols 'Last Name' 'First Name')")
    parser.add_argument("--student_cell", default="B4", help="Cell to write member emails (default: B4)")
    parser.add_argument("--grade_cell", default="B11", help="Cell reserved for grade (default: B11)")
    parser.add_argument("--template", default=None, help="Optional path to an .xlsx template to copy")
    parser.add_argument("--sheet_name", default=None, help="Optional sheet name to target (if template has multiple)")
    parser.add_argument("--overwrite", action="store_true", help="Overwrite existing files")

    args = parser.parse_args()

    try:
        paths = generate_group_sheets(
            master_path=args.master,
            output_folder=args.outdir,
            email_col=args.email_col,
            group_col=args.group_col,
            name_cols=args.name_cols,
            student_cell=args.student_cell,
            grade_cell=args.grade_cell,
            template_path=args.template,
            sheet_name=args.sheet_name,
            overwrite=args.overwrite,
        )
        print(f"Generated {len(paths)} files:")
        for p in paths:
            print(" -", p)
    except Exception as e:
        print(f"Error: {e}")
